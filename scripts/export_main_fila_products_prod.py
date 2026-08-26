"""
斐乐商品数据导出脚本（生产环境 - 精简版）
- 商品接口: SP_M_SPZSJ_QUERY (不带子表, PAGE_SIZE=100)
- 字典接口: MDM_MODEL_BASE_M_DATA_DIC (批量查询)
- 商品默认按最近 2 年 LAST_UPDATE_DATE 拉取（运行日动态计算）
- 输出: fila_products_brief_prod.xlsx (29列精简版，含集团品牌/季节/配货季)
"""

import hashlib
import json
import time
import uuid
import logging
from datetime import datetime, timedelta, timezone
from typing import Dict, List

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============== 配置 ==============

# 生产环境地址
PRODUCT_API_URL = (
    "http://aop.antaapi.com/ASH241120162739307"
    "/anta/mdm/demdm-api/open/api/v2/selectApi/SP_M_SPZSJ_QUERY/1.0.0"
)
DICT_API_URL = (
    "http://aop.antaapi.com/ASH241120162739307"
    "/anta/mdm/demdm-api/open/api/v2/selectApi/MDM_MODEL_BASE_M_DATA_DIC/1.0.0"
)

# 生产环境密钥
APP_KEY = "ASH24112014534835"
APP_SECRET = "8af5b725eac941b1ad614034e0c4f54d"

# 斐乐品牌编码
FILA_FASHGRD = ["1101", "1102", "1103", "1104"]

# 商品 LAST_UPDATE_DATE 默认回溯年数（相对运行日）
PRODUCT_UPDATE_YEARS = 2
TZ_SHANGHAI = timezone(timedelta(hours=8))

# 拉取参数
PRODUCT_PAGE_SIZE = 100
DICT_PAGE_SIZE = 300
REQUEST_INTERVAL = 0.05

# 输出文件
OUTPUT_FILE = "fila_products_brief_prod.xlsx"

logger = logging.getLogger(__name__)


# ============== 工具函数 ==============

def generate_sign(app_key: str, secret: str, timestamp: str) -> str:
    raw = app_key + secret + timestamp
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def build_headers() -> dict:
    timestamp = str(int(time.time() * 1000))
    sign = generate_sign(APP_KEY, APP_SECRET, timestamp)
    return {
        "op-api-key": APP_KEY,
        "op-api-timestamp": timestamp,
        "op-api-sign": sign,
        "Content-Type": "application/json",
    }


def safe_get(obj: dict, key: str, default=""):
    v = obj.get(key)
    return v if v is not None else default


def _fmt_api_time(dt: datetime) -> str:
    """格式化为 API 要求的 +08:00 时间串。"""
    return dt.strftime("%Y-%m-%dT%H:%M:%S") + "+08:00"


def product_update_time_range(
    years: int = PRODUCT_UPDATE_YEARS,
) -> tuple[str, str]:
    """按运行日计算商品拉取时间窗：最近 years 年至今日末。"""
    now = datetime.now(TZ_SHANGHAI)
    end = now.replace(hour=23, minute=59, second=59, microsecond=0)
    start = (now - timedelta(days=365 * years)).replace(
        hour=0, minute=0, second=0, microsecond=0,
    )
    return _fmt_api_time(start), _fmt_api_time(end)


# ============== 字典拉取 ==============

def fetch_dict_batch(
    category_codes: List[str],
    max_pages: int = 50,
    max_retries: int = 3,
) -> List[dict]:
    """批量拉取多个 CATEGORY_CODE 的字典数据（自动翻页）"""
    all_data = []
    last_id = None
    page = 0

    while page < max_pages:
        page += 1
        logger.info(f"[字典] 第 {page} 页, codes={category_codes}, last_id={last_id}")

        body = {
            "FORM_CODE": "ALL",
            "UUID": uuid.uuid4().hex.upper(),
            "PAGE": str(page),
            "PAGE_SIZE": str(DICT_PAGE_SIZE),
            "DATA": {
                "CATEGORY_CODE": category_codes,
                "LAST_UPDATE_DATE_START": "2020-01-01T00:00:00+08:00",
                "LAST_UPDATE_DATE_END": "2026-12-31T23:59:59+08:00",
            },
        }
        if last_id is not None:
            body["LAST_ID"] = last_id

        retries = 0
        result = None
        while retries < max_retries:
            try:
                resp = requests.post(DICT_API_URL, headers=build_headers(), json=body, timeout=30)
                resp.raise_for_status()
                result = resp.json()
                break
            except requests.RequestException as e:
                retries += 1
                logger.warning(f"[字典] 重试 {retries}/{max_retries}: {e}")
                if retries >= max_retries:
                    return all_data
                time.sleep(2 ** retries)

        if result is None:
            break
        if result.get("RESULT", "") != "S":
            logger.error(f"[字典] 失败: {result.get('MESSAGE', '')}")
            break

        data_list = result.get("DATA", [])
        if not data_list:
            break

        all_data.extend(data_list)
        logger.info(f"[字典] 本页 {len(data_list)} 条, 累计 {len(all_data)} 条")

        if result.get("LAST_PAGE", True):
            break

        last_id = (data_list[-1].get("ID") or data_list[-1].get("id", ""))
        if not last_id:
            break
        time.sleep(REQUEST_INTERVAL)

    return all_data


def build_dict_lookup() -> Dict[str, Dict[str, str]]:
    """从生产环境 API 拉取所有需要的字典，构建 code→text 对照表"""
    # 精简版需要的字典编码（对应26列中需要翻译的字段）
    dict_codes = [
        "SPART",          # 大类
        "ZCATEGNOTES",    # 中类
        "ZCATEGORY",      # 小类
        "ZSERIES",        # 系列
        "SUBSERIES",      # 子系列
        "ZSEX",           # 性别
        "ZAGEGROUP",      # 年龄段
        "ZPRICERANGE",    # 价格带
        "FASHGRD",        # 品牌
        "ZSEASON",        # 季节
        "MTART",          # 物料类型
        "MATKL",          # 物料组
        "MSTAE",          # 淘汰状态
        "MEINS",          # 基本计量单位
        "ZCJBH",          # 场景编号
        "ZCMFWLX",        # 概念品牌
        "ZDDLX",          # 下单类型
        "ZXDLX",          # 订单类型
        "ZYWLX",          # 业务类型
        "ZPRODUCTMIXC",   # 产品类别
        "ZPMGJ",          # 产品经理
        "ZCATEGORY_TAG",  # 品类标签
    ]

    logger.info(f"[字典] 从生产 API 批量拉取 {len(dict_codes)} 个品类: {dict_codes}")
    raw_data = fetch_dict_batch(dict_codes)

    # 按 CATEGORY_CODE 分组
    lookup: Dict[str, Dict[str, str]] = {c: {} for c in dict_codes}
    for item in raw_data:
        code = item.get("CATEGORY_CODE", "")
        cv = item.get("CHARACTERISTIC_VALUE", "")
        zh = next((t.get("CHARACTERISTIC_TEXT", "") for t in item.get("MULTI_LANG", [])
                   if t.get("LANG") == "zh-CN"), "")
        if code in lookup and cv:
            lookup[code][cv] = zh or cv

    for code in dict_codes:
        logger.info(f"[字典] {code}: {len(lookup[code])} 条")

    return lookup


# ============== 商品拉取 ==============

def fetch_all_products(
    last_update_start: str,
    last_update_end: str,
    fashgrd: list,
    max_pages: int = 5000,
    max_retries: int = 3,
) -> list:
    """分页拉取全部商品主数据（自动翻页，不带子表）"""
    all_data = []
    last_id = None
    page = 0

    while page < max_pages:
        page += 1
        logger.info(f"[商品] 第 {page} 页, last_id={last_id}")

        body = {
            "QUERY_CHILD_TABLE": [],
            "FORM_CODE": "SP_M_SPZSJ_FILA",
            "UUID": uuid.uuid4().hex.upper(),
            "PAGE": "1",
            "PAGE_SIZE": str(PRODUCT_PAGE_SIZE),
            "DATA": {
                "LAST_UPDATE_DATE_START": last_update_start,
                "LAST_UPDATE_DATE_END": last_update_end,
                "FASHGRD": fashgrd,
                "MATNR": "",
                "MTART": "",
                "SPART": "",
            },
        }
        if last_id is not None:
            body["LAST_ID"] = str(last_id)

        retries = 0
        result = None
        while retries < max_retries:
            try:
                resp = requests.post(PRODUCT_API_URL, headers=build_headers(), json=body, timeout=30)
                resp.raise_for_status()
                result = resp.json()
                break
            except requests.RequestException as e:
                retries += 1
                logger.warning(f"[商品] 重试 {retries}/{max_retries}: {e}")
                if retries >= max_retries:
                    return all_data
                time.sleep(2 ** retries)

        if result is None:
            break
        if result.get("RESULT", "") != "S":
            logger.error(f"[商品] 失败: {result.get('MESSAGE', '')}")
            break

        data_list = result.get("DATA", [])
        if not data_list:
            break

        all_data.extend(data_list)
        logger.info(f"[商品] 本页 {len(data_list)} 条, 累计 {len(all_data)} 条")

        if result.get("LAST_PAGE", True):
            break

        last_id = (data_list[-1].get("ID") or data_list[-1].get("id", ""))
        if not last_id:
            break
        time.sleep(REQUEST_INTERVAL)

    return all_data


# ============== 翻译 & 组装（精简26列） ==============

def translate(code: str, lookup: Dict[str, Dict[str, str]], field: str) -> str:
    """将编码翻译为中文文本"""
    if not code:
        return ""
    dict_map = lookup.get(field, {})
    return dict_map.get(str(code), str(code))


# 精简版29列表头
BRIEF_COLUMNS = [
    "货号",
    "款号",
    "产品名称",
    "集团品牌",
    "大类",
    "中类",
    "小类",
    "系列",
    "子系列",
    "性别",
    "针梭织",
    "上下装",
    "POP-穿着明星",
    "颜色",
    "颜色文本",
    "颜色色系",
    "版型/楦型",
    "面材料/成分",
    "季节",
    "季节属性",
    "配货季",
    "场景",
    "场景_标签",
    "子场景",
    "颜色风格",
    "风格",
    "图案位置",
    "图案特征",
    "领型裤脚口裙型",
]


def build_brief_rows(products: list, lookup: Dict[str, Dict[str, str]]) -> List[dict]:
    """将商品数据翻译并组装为精简版 Excel 行（29列）"""
    rows = []

    for item in products:
        ml_list = item.get("MULTI_LANG", [])
        zh = next((m for m in ml_list if m.get("LANG") == "zh-CN"), {})

        row = {}

        row["货号"] = safe_get(item, "MATNR")
        row["款号"] = safe_get(item, "ZDESIGNNUM")
        row["产品名称"] = safe_get(zh, "MAKTX")
        row["集团品牌"] = translate(item.get("FASHGRD", ""), lookup, "FASHGRD")
        row["大类"] = translate(item.get("SPART", ""), lookup, "SPART")
        row["中类"] = translate(item.get("ZCATEGNOTES", ""), lookup, "ZCATEGNOTES")
        row["小类"] = translate(item.get("ZCATEGORY", ""), lookup, "ZCATEGORY")
        row["系列"] = translate(item.get("ZSERIES", ""), lookup, "ZSERIES")
        row["子系列"] = translate(item.get("SUBSERIES", ""), lookup, "SUBSERIES")
        row["性别"] = translate(item.get("ZSEX", ""), lookup, "ZSEX")
        row["针梭织"] = ""  # 不在主表中
        row["上下装"] = safe_get(item, "ZCLOTHSSET")
        row["POP-穿着明星"] = ""  # 不在主表中
        row["颜色"] = safe_get(zh, "ZCOLOR")
        row["颜色文本"] = safe_get(zh, "ZCOLOR_TX")
        row["颜色色系"] = ""  # 不在主表中
        row["版型/楦型"] = ""  # 不在主表中
        row["面材料/成分"] = safe_get(zh, "ZSUREFACE")
        row["季节"] = translate(item.get("SEASON", ""), lookup, "ZSEASON")
        row["季节属性"] = safe_get(item, "J_AMSEA")
        row["配货季"] = safe_get(item, "ZDSEASON")
        row["场景"] = safe_get(item, "ZCJBH")
        row["场景_标签"] = safe_get(zh, "ZSCENE")
        row["子场景"] = ""  # 不在主表中
        row["颜色风格"] = ""  # 不在主表中
        row["风格"] = ""  # 不在主表中
        row["图案位置"] = ""  # 不在主表中
        row["图案特征"] = ""  # 不在主表中
        row["领型裤脚口裙型"] = safe_get(zh, "ZCARDIGANMETHOD")

        rows.append(row)

    return rows


# ============== Excel 输出 ==============

def write_excel(rows: List[dict], filepath: str):
    """将数据写入 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品_斐乐_精简"

    # 样式
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, col_name in enumerate(BRIEF_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写数据
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(BRIEF_COLUMNS, 1):
            value = row_data.get(col_name, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动列宽（采样前100行）
    for col_idx, col_name in enumerate(BRIEF_COLUMNS, 1):
        max_width = len(str(col_name)) * 2
        for row_idx in range(2, min(len(rows) + 2, 102)):
            val = rows[row_idx - 2].get(col_name, "")
            width = len(str(val)) * 2 if any('\u4e00' <= c <= '\u9fff' for c in str(val)) else len(str(val))
            max_width = max(max_width, width)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 4, 40)

    wb.save(filepath)
    logger.info(f"Excel 已保存: {filepath} ({len(rows)} 行)")


# ============== 主流程 ==============

def run_export(output_file: str = OUTPUT_FILE, years: int = PRODUCT_UPDATE_YEARS) -> int:
    """执行字典拉取 -> 商品拉取 -> 翻译组装 -> 写 Excel 的完整流程。

    供 scripts/daily_download_product_tables.py 复用；返回写入行数。
    """
    logger.info("=" * 60)
    logger.info("斐乐商品数据导出（生产环境-精简版）开始")
    logger.info("=" * 60)

    # Step 1: 拉取字典（从生产 API）
    logger.info(">>> Step 1: 从生产环境拉取字典翻译表")
    lookup = build_dict_lookup()
    total_entries = sum(len(v) for v in lookup.values())
    logger.info(f"字典拉取完成: {len(lookup)} 个品类, 共 {total_entries} 条")

    # Step 2: 拉取商品（生产环境，默认最近 years 年更新）
    time_start, time_end = product_update_time_range(years=years)
    logger.info(
        f">>> Step 2: 从生产环境拉取斐乐商品 "
        f"(最近 {years} 年, {time_start} ~ {time_end})",
    )
    products = fetch_all_products(
        last_update_start=time_start,
        last_update_end=time_end,
        fashgrd=FILA_FASHGRD,
    )
    logger.info(f"商品拉取完成: {len(products)} 条")

    if not products:
        logger.warning("未拉取到商品数据！")
        return 0

    # Step 3: 翻译并组装为精简版行
    logger.info(f">>> Step 3: 翻译编码并组装精简版（{len(BRIEF_COLUMNS)}列）")
    rows = build_brief_rows(products, lookup)
    logger.info(f"组装完成: {len(rows)} 行")

    # Step 4: 输出 Excel
    logger.info(">>> Step 4: 生成 Excel")
    write_excel(rows, output_file)

    # 统计
    filled_cols = sum(1 for c in BRIEF_COLUMNS if any(r.get(c) for r in rows))
    logger.info(f"{'=' * 60}")
    logger.info(f"导出完成: {output_file}")
    logger.info(f"总行数: {len(rows)}, 总列数: {len(BRIEF_COLUMNS)}, 有数据列: {filled_cols}")
    logger.info(f"{'=' * 60}")
    return len(rows)


def setup_logging() -> None:
    """独立运行时配置日志（文件 + 控制台）。被复用方 import 时不调用，避免污染其日志。"""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler("export_prod.log", encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def main():
    setup_logging()
    run_export()


if __name__ == "__main__":
    main()
