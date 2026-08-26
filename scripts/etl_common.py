"""FILA 离线 ETL 共用：读表、货号解析、角色/日志（对齐 descente_agent_html/scripts）。"""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import defaultdict

# product_sku.csv 等表中存在单字段 ~16MB 的内嵌大字段（blob/图片等），
# 超出 csv 模块默认 131072 字节上限，这里放宽到平台上限。
csv.field_size_limit(sys.maxsize)
from dataclasses import dataclass, field
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any, DefaultDict, Iterator

import yaml


def _parse_iso_datetime(raw: Any) -> datetime | None:
    """Parse an ISO 8601 datetime string; returns None on failure."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Strip milliseconds
    s = re.sub(r'\.\d+', '', s)
    for fmt in (
        '%Y-%m-%dT%H:%M:%S%z',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

from scripts._project_paths import load_paths

_PATHS = load_paths()
ROOT = _PATHS["root"]
REPO_ROOT = _PATHS["repo_root"]

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(_PATHS["tools_dir"]) not in sys.path:
    sys.path.insert(0, str(_PATHS["tools_dir"]))

from backend.models import normalize_season
from backend.intent.sku_attributes import (
    extract_coverage,
    extract_is_intimate,
    extract_layer,
    extract_length_class,
    extract_scene_domain,
    is_swimwear,
    normalize_modeling,
)
from backend.intent.color_series_mapper import map_color_to_series_list

BRAND = "FILA"
# 上架时间下限：只保留 2023-01-01 及以后上架的商品
MIN_UP_TIME = datetime(2023, 1, 1)
_ATTR_COLOR_RE = re.compile(r"颜色:([^;]+);尺码:")

# descent 复刻：brand_line 由 id_brand 映射（与 build_catalog._FILA_BRAND_IDS 一致）
_FILA_BRAND_LINE_MAP: dict[str, str] = {
    "1": "FILA",
    "17": "FILA KIDS",
    "21": "FILA FUSION",
    "10": "FILA联名",
}


def _safe_float(val: Any) -> float:
    if val is None:
        return 0.0
    try:
        v = float(val)
        return v if 0 < v < 100000 else 0.0
    except (TypeError, ValueError):
        return 0.0


def _safe_int(val: Any) -> int:
    if val is None:
        return 0
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return 0


def merge_features(pro_info: str, pro_content: str) -> str:
    """合并 pro_info + pro_content，逗号分隔去重保序（descent merge_features 同逻辑）。"""
    parts: list[str] = []
    for src in (pro_info, pro_content):
        if src:
            parts.extend(p.strip() for p in str(src).split(",") if p.strip())
    seen: set[str] = set()
    result: list[str] = []
    for p in parts:
        if p and p not in seen:
            seen.add(p)
            result.append(p)
    return ", ".join(result)


def build_descent_extra_fields(
    master: dict[str, Any],
    ext: dict[str, Any],
    color_attr_rows: list[dict[str, Any]],
    sku_count: int,
) -> dict[str, Any]:
    """从 product_master/product_master_ext/product_attr 派生 descent 用到的、fila 原缺字段。

    纯函数：输入原始表行，输出可直接并入 sku record 的字段 dict。
    color_attr_rows 为该 goods 的 id_pac=1 颜色属性行（含 image_url）。
    """

    def _ts(v: Any) -> str:
        return text_or_empty(v)

    color_images = [
        {
            "color": _ts(r.get("attr_name")),
            "image_url": _ts(r.get("image_url")),
        }
        for r in (color_attr_rows or [])
        if _ts(r.get("image_url"))
    ]
    return {
        "product_name_short": _ts(master.get("pro_name")),
        "goods_sn": _ts(master.get("id_alias")),
        "brand_line": _FILA_BRAND_LINE_MAP.get(_ts(master.get("id_brand")), ""),
        "market_price": _safe_float(master.get("market_price")),
        "min_price": _safe_float(master.get("min_price")),
        "max_price": _safe_float(master.get("max_price")),
        "year": _ts(ext.get("year")),
        "category": _ts(ext.get("cat_alias")),
        "length": _ts(ext.get("length")),
        "technology": _ts(ext.get("technology")),
        "features": merge_features(
            _ts(master.get("pro_info")), _ts(master.get("pro_content"))
        ),
        "selling_point_label": _ts(master.get("selling_point_label")),
        "keyword": _ts(master.get("keyword")),
        "color_images": json.dumps(color_images, ensure_ascii=False) if color_images else "",
        "video_url": _ts(master.get("video")),
        "onsell": _safe_int(master.get("onsell")),
        "sales": _safe_int(master.get("sales")),
        "sales_week": _safe_int(master.get("sales_week")),
        "sales_month": _safe_int(master.get("sales_month")),
        "w_order": _safe_int(master.get("w_order")),
        "sku_count": int(sku_count or 0),
    }



def load_cfg() -> dict[str, Any]:
    with (ROOT / "config.yaml").open(encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def product_dir(cfg: dict[str, Any] | None = None) -> Path:
    paths = (cfg or load_cfg()).get("paths") or {}
    return ROOT / str(paths.get("product_dir", "data/tables"))


def processed_dir(cfg: dict[str, Any] | None = None) -> Path:
    paths = (cfg or load_cfg()).get("paths") or {}
    return ROOT / str(paths.get("processed_dir", "data/processed"))


def logs_dir(cfg: dict[str, Any] | None = None) -> Path:
    paths = (cfg or load_cfg()).get("paths") or {}
    return ROOT / str(paths.get("logs_dir", "data/logs"))


def reports_dir() -> Path:
    p = ROOT / "data" / "reports"
    p.mkdir(parents=True, exist_ok=True)
    return p


def load_csv(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def norm_id(raw: Any) -> int | None:
    if raw is None:
        return None
    s = str(raw).strip().strip('"').lstrip("'").strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def norm_id_pa(val: object) -> str:
    if val is None or val is False:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    try:
        num = float(s)
        if num == int(num):
            return str(int(num))
    except ValueError:
        pass
    return s


def text_or_empty(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def is_onsell(raw: Any) -> bool:
    """product_master.onsell：1 或 2 表示在售/可售（0=下架，3/4=其它状态）。

    None / 空 / 非法一律视为不在售，配合 ``is_up_time_ok`` 实现「onsell 与
    up_time 均不能为空」的严格过滤。
    """
    if raw is None:
        return False
    s = str(raw).strip()
    if not s:
        return False
    try:
        return int(float(s)) in (1, 2)
    except ValueError:
        return s in ("1", "2")


def is_legacy_sku_id(sku_id: str) -> bool:
    """识别老款/电商款 sku_id（数字开头，如 162217109-3、112128861-1、0028728）。

    新款 FILA 货号均以字母开头（A11M411206FBU、A1EU621231FWT、F1EU629038FLG）。
    用于在 ETL 入口过滤掉上游 product_attr 中遗留的老款颜色货号，统一 sku_id 体系。
    """
    if not sku_id:
        return False
    return sku_id[0].isdigit()


def normalize_spu_id(spu: str) -> str:
    """剥离 SPU 开头的 4 位年份前缀（如 2025A13W447361F -> A13W447361F）。

    上游 product_master.id_alias 部分新款带年份前缀（2025/2026...），
    剥离后使 spu_id 与 spu_to_skus.json 的 key 统一为标准 SPU 形式。
    仅当前 4 位全为数字且第 5 位为字母时才剥离，避免误伤纯数字 id_alias。
    """
    if len(spu) > 4 and spu[:4].isdigit() and spu[4:5].isalpha():
        return spu[4:]
    return spu


def is_up_time_ok(raw: Any) -> bool:
    """product_master.up_time：上架时间 >= MIN_UP_TIME（2023-01-01）。"""
    dt = _parse_iso_datetime(raw)
    if dt is None:
        # 无上架时间字段时视为不符合，避免引入脏数据
        return False
    # 去掉时区信息以便与 naive datetime 比较
    if dt.tzinfo is not None:
        dt = dt.replace(tzinfo=None)
    return dt >= MIN_UP_TIME


def normalize_up_time(raw: Any) -> str:
    """归一化 product_master.up_time 为 'YYYY-MM-DD HH:MM:SS' 字符串。

    up_time 的单一真源：ES date 字段直接解析该字符串，Milvus 侧由
    ``up_time_to_epoch`` 再转成 epoch 秒。解析失败返回 ''。
    """
    dt = _parse_iso_datetime(raw)
    if dt is None:
        return ""
    if dt.tzinfo is not None:
        dt = dt.replace(tzinfo=None)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def up_time_to_epoch(raw: Any) -> int:
    """up_time -> epoch 秒（INT64），供 Milvus 标量倒排/排序用；失败返回 0。

    按 UTC 计算以保证不同机器时区下结果一致（源数据 up_time 无时区）。
    """
    dt = _parse_iso_datetime(raw)
    if dt is None:
        return 0
    if dt.tzinfo is not None:
        dt = dt.replace(tzinfo=None)
    return int(dt.replace(tzinfo=timezone.utc).timestamp())


def first_non_empty(*vals: Any) -> str | None:
    for val in vals:
        txt = text_or_empty(val)
        if txt:
            return txt
    return None


def extract_color_name(attr_name: str) -> str:
    m = _ATTR_COLOR_RE.search(attr_name or "")
    if m:
        return m.group(1).strip()
    return text_or_empty(attr_name)


def fila_category_l2(ext: dict[str, str]) -> str:
    raw = text_or_empty(
        first_non_empty(ext.get("middle_class"), ext.get("cat_alias")) or "",
    )
    return normalize_category_l2(raw)


_CATEGORY_L2_MERGE: dict[str, str] | None = None


def _load_category_l2_merge() -> dict[str, str]:
    global _CATEGORY_L2_MERGE
    if _CATEGORY_L2_MERGE is not None:
        return _CATEGORY_L2_MERGE
    merge_path = Path(__file__).resolve().parent.parent / "backend" / "intent" / "dictionaries" / "category_l2_merge.yaml"
    if merge_path.is_file():
        with merge_path.open(encoding="utf-8") as f:
            _CATEGORY_L2_MERGE = yaml.safe_load(f) or {}
    else:
        _CATEGORY_L2_MERGE = {}
    return _CATEGORY_L2_MERGE


def normalize_category_l2(raw: str) -> str:
    """将原始细分中类归一化为合并后的标准中类。"""
    if not raw:
        return raw
    merge = _load_category_l2_merge()
    return merge.get(raw, raw)


_EXCLUDED_CATEGORIES: set[str] | None = None


def _load_excluded_categories() -> set[str]:
    """加载 non_clothing_exclusion.yaml 中不参与搭配的中类集合。"""
    global _EXCLUDED_CATEGORIES
    if _EXCLUDED_CATEGORIES is not None:
        return _EXCLUDED_CATEGORIES
    excl_path = (
        Path(__file__).resolve().parent.parent
        / "backend" / "intent" / "dictionaries" / "non_clothing_exclusion.yaml"
    )
    if excl_path.is_file():
        with excl_path.open(encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        _EXCLUDED_CATEGORIES = set(
            data.get("non_clothing", []) + data.get("intimate_swimwear", [])
        )
    else:
        _EXCLUDED_CATEGORIES = set()
    return _EXCLUDED_CATEGORIES


def is_excluded_from_pairing(category_l2: str) -> bool:
    """判断中类是否被排除出搭配（非服饰/内衣泳衣）。"""
    return category_l2 in _load_excluded_categories()


def split_tags(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[,，;；/|]", raw)
    return [p.strip() for p in parts if p.strip()]


_SX_UP_DOWN_MAP: dict[str, str] = {
    "SX01": "上装",
    "SX02": "下装",
    "SX05": "连衣裙",
}

CATEGORY_L2_UP_DOWN: dict[str, str] = {
    # ── 上装 ──────────────────────────────────────────────────────────
    "短袖T": "上装", "短T": "上装", "短T类": "上装", "短袖T恤": "上装", "T恤": "上装",
    "长袖T": "上装", "长袖T恤": "上装", "长T": "上装",
    "棉质短袖T恤": "上装", "棉质长袖T恤": "上装",
    "速干短袖T恤": "上装", "速干长袖T恤": "上装",
    "羊毛短袖T恤": "上装", "羊毛长袖T恤": "上装",
    "针织运动上衣": "上装", "针织短袖衫": "上装", "短袖针织衫": "上装",
    "短袖针织上衣": "上装",
    "针织套头衫": "上装", "长袖针织衫": "上装", "针织上衣": "上装",
    "针织背心": "上装", "针织运动套": "上装", "针织连帽套头衫": "上装",
    "短袖POLO": "上装", "短袖POLO衫": "上装", "长袖POLO": "上装",
    "羊毛POLO衫": "上装",
    "套头衫": "上装", "套头卫衣": "上装", "连帽卫衣": "上装", "卫衣": "上装",
    "梭织上衣": "上装", "梭织运动上衣": "上装",
    "短袖编织衫": "上装", "编织衫": "上装", "编织开衫": "上装",
    "短袖衬衫": "上装", "长袖衬衫": "上装", "衬衫": "上装",
    "短袖衬衣": "上装", "长袖衬衣": "上装",
    "背心": "上装", "运动背心": "上装",
    "上衣": "上装", "短装": "上装",
    "内搭": "上装", "内搭类": "上装", "内层长袖上装": "上装", "内衣类": "上装",
    "运动内衣": "上装", "内着类": "上装",
    "外套": "上装", "外套类": "上装",
    "冲锋衣": "上装", "三合一夹克": "上装", "单层冲锋衣": "上装",
    "软壳茄克": "上装", "软壳夹克": "上装",
    "防风茄克": "上装", "防风夹克": "上装", "防水夹克": "上装",
    "梭织薄外套": "上装", "梭织厚外套": "上装", "梭织外套": "上装",
    "抓绒衫": "上装", "摇粒绒运动上衣": "上装", "丝光绒运动上衣": "上装",
    "保暖棉服": "上装", "棉服": "上装", "棉服大衣": "上装",
    "棉羽": "上装", "棉羽类": "上装",
    "羽绒夹克": "上装", "羽绒服": "上装", "羽绒大衣": "上装",
    "羽绒马甲": "上装", "羽绒马夹": "上装",
    "常规短羽绒服": "上装", "中长羽绒服": "上装",
    "滑雪服": "上装", "滑雪茄克": "上装",
    "马甲": "上装", "梭织马甲": "上装",
    "防晒服": "上装",
    "毛衣": "上装",
    "BRA": "上装",
    "冬装": "上装",
    # ── 下装 ──────────────────────────────────────────────────────────
    "长裤": "下装", "梭织长裤": "下装", "针织长裤": "下装",
    "梭织裤": "下装", "针织裤": "下装",
    "梭织运动长裤": "下装", "针织运动长裤": "下装",
    "梭织五分裤": "下装", "针织五分裤": "下装",
    "梭织短裤": "下装", "针织短裤": "下装",
    "梭织七分裤": "下装", "针织七分裤": "下装",
    "中裤": "下装", "短裤": "下装", "短裤类": "下装",
    "速干长裤": "下装", "速干短裤": "下装",
    "内层长裤": "下装", "针织打底裤": "下装",
    "滑雪长裤": "下装", "冲锋裤": "下装", "软壳长裤": "下装",
    "半身裙": "下装", "梭织半裙": "下装", "针织半裙": "下装",
    "梭织裤裙": "下装", "裙装": "下装",
    "下装类": "下装",
    "休闲长裤": "下装",
    "紧身裤": "下装",
    # ── 连衣裙 ────────────────────────────────────────────────────────
    "连衣裙": "连衣裙", "针织连衣裙": "连衣裙", "梭织连衣裙": "连衣裙",
    # ── 套装（上下装组合，不单独分类） ───────────────────────────────
    "套装": "套装", "两件套": "套装", "梭织两件套": "套装",
    "篮球比赛套": "套装", "足球比赛套": "套装",
    # ── 合并后新增中类名 → 角色映射 ─────────────────────────────────
    "裤裙": "下装",
    "板鞋": "鞋", "老爹鞋": "鞋", "跑鞋": "鞋", "运动鞋": "鞋",
    "潮鞋": "鞋", "户外鞋": "鞋", "休闲鞋": "鞋", "帆布鞋": "鞋",
    "凉鞋": "鞋", "拖鞋": "鞋", "网球鞋": "鞋", "高尔夫鞋": "鞋",
    "儿童鞋": "鞋",
    "帽子": "配饰", "袜子": "配饰", "包": "配饰",
}


_SKU_GENDER_CODE: dict[str, str] = {
    "M": "男",
    "W": "女",
    "B": "男童",
    "G": "女童",
}


def infer_gender_from_sku(sku_id: str, title: str) -> str:
    """从 SKU 编码第4位或商品标题推断性别，作为 ext.sex 为空时的回退。"""
    # 标准编码: 品牌(1) + 季(2) + 性别码(1), e.g. F11M..., A12W...
    if len(sku_id) >= 4 and sku_id[0] in "FATK":
        g = _SKU_GENDER_CODE.get(sku_id[3])
        if g:
            return g
        if sku_id[3] == "U":
            return "中性"
    # 从标题推断
    if title:
        if "男女" in title:
            return "中性"
        for kw, val in (("女童", "女童"), ("男童", "男童"),
                        ("女小童", "女童"), ("男小童", "男童"),
                        ("女大童", "女童"), ("男大童", "男童"),
                        ("女子", "女"), ("男子", "男"),
                        ("女士", "女"), ("男士", "男")):
            if kw in title:
                return val
    return ""


_GENDER_CANONICAL_LIST: frozenset[str] = frozenset({"男", "女", "男童", "女童", "儿童"})

_GENDER_MULTI_MAP: dict[str, list[str]] = {
    "男": ["男"], "男士": ["男"], "男生": ["男"], "男性": ["男"], "男装": ["男"],
    "女": ["女"], "女士": ["女"], "女生": ["女"], "女性": ["女"], "女装": ["女"],
    "男童": ["男童"], "男孩": ["男童"], "男宝": ["男童"], "小男生": ["男童"],
    "女童": ["女童"], "女孩": ["女童"], "女宝": ["女童"], "小女生": ["女童"],
    "儿童": ["儿童"], "童装": ["儿童"], "小朋友": ["儿童"], "孩子": ["儿童"],
    "宝宝": ["儿童"], "小孩": ["儿童"],
    "男女同款": ["男", "女"],
    "男女": ["男", "女"],
    "男/女": ["男", "女"],
    "男女随机": ["男", "女"],
    "男女童": ["男童", "女童"],
    "中性": ["男", "女"],
    "中": ["男", "女"],
    "中性款": ["男", "女"],
    "通用": ["男", "女"],
    "不分性别": ["男", "女"],
    "UNISEX": ["男", "女"],
    "MAN": ["男"],
    "WOMAN": ["女"],
}


def normalize_gender_to_list(
    raw: str,
    title: str = "",
    sku_id: str = "",
) -> list[str]:
    """将原始 sex 字段归一化为标准 gender 列表，元素来自 {男,女,男童,女童,儿童}。

    多值/中性表述展开为多元素列表，便于下游 ES/Milvus 数组字段精确匹配。
    无法识别时回退到 SKU 编码/标题推断。
    """
    s = (raw or "").strip()
    if s:
        mapped = _GENDER_MULTI_MAP.get(s)
        if mapped:
            return list(mapped)
        if s in _GENDER_CANONICAL_LIST:
            return [s]
    inferred = infer_gender_from_sku(sku_id, title)
    if inferred:
        if inferred == "中性":
            return ["男", "女"]
        return [inferred]
    return []


# 童装年龄段（源数据 product_master_ext.age 列），与性别正交的独立维度。
# 通码 = 不分年龄段（同款覆盖小童~中大童）。
_AGE_CANONICAL: frozenset[str] = frozenset({"小童", "中大童", "婴幼童", "通码"})


def normalize_age(value: object) -> str:
    """归一化原始 age 字段为标准年龄段字符串。

    源数据 age 列取值：小童 / 中大童 / 婴幼童 / 通码。
    空值或噪声（如 "33"）返回 ""。注意这是 age 而非 age_group——
    直接保留源端取值，不重分桶。
    """
    s = text_or_empty(value).strip()
    if not s:
        return ""
    if s in _AGE_CANONICAL:
        return s
    # 容错：去空白/全角后命中
    s2 = s.replace(" ", "").replace("　", "")
    if s2 in _AGE_CANONICAL:
        return s2
    return ""


# 童装年龄段 title→标准桶映射（与 ages.yaml / backend.models.normalize_age 别名一致）。
# 按特异性排序：长串/婴幼优先，避免「婴幼童」被「幼童」、「中大童」被「大童」误截。
_AGE_TITLE_MAP: list[tuple[str, str]] = [
    ("婴幼童", "婴幼童"),
    ("婴儿", "婴幼童"),
    ("婴幼", "婴幼童"),
    ("幼童", "婴幼童"),
    ("中大童", "中大童"),
    ("大童", "中大童"),
    ("中童", "中大童"),
    ("小童", "小童"),
    ("通码", "通码"),
]


def infer_age_from_title(title: str) -> str:
    """从商品名称推断童装年龄段，作为源 age 列为空时的兜底。

    只匹配童装年龄段关键词，故天然只对童装生效——成人款标题不含这些词，返回 ""。
    映射与 ages.yaml / backend.models.normalize_age 的别名归一一致：
    大童/中童→中大童，婴儿/婴幼/幼童→婴幼童。
    """
    if not title:
        return ""
    for kw, canonical in _AGE_TITLE_MAP:
        if kw in title:
            return canonical
    return ""


def infer_up_down_from_title(title: str) -> str:
    """从商品名称推断上下装，作为 cat_l2 映射找不到时的最终回退。"""
    if not title:
        return ""
    if "连衣裙" in title:
        return "连衣裙"
    for kw in ("裤子", "长裤", "短裤", "半裙", "半身裙", "裤裙", "打底裤",
               "紧身裤", "五分裤", "七分裤", "中裤"):
        if kw in title:
            return "下装"
    for kw in ("T恤", "POLO", "卫衣", "外套", "夹克", "茄克", "大衣", "棉服",
               "马甲", "背心", "衬衫", "衬衣", "毛衣", "开衫", "针织衫",
               "内搭", "上衣", "BRA", "冲锋衣", "羽绒服", "防风", "抓绒",
               "摇粒绒", "套头衫", "连帽衫", "防晒服", "风衣"):
        if kw in title:
            return "上装"
    return ""


def infer_role(
    ext: dict[str, str],
    title: str,
    up_down_extra: str = "",
    sub_cat: str = "",
) -> str:
    """推断角色 role。

    sub_cat 为小类（cat_l3 / short_category），仅在上述强信号均未命中、即将落
    unknown 时兜底：套装/两件套等 up_down=「套装」落不进 ud_map，但小类往往标了
    细分单品（「梭织上衣」→top、「梭织裤」→bottoms），据此补判角色。
    """
    ud = text_or_empty(up_down_extra) or text_or_empty(ext.get("up_down"))
    cat = text_or_empty(ext.get("cat_alias") or ext.get("cat"))
    ct = text_or_empty(ext.get("cat_type"))
    ttl = title or ""
    ud_map = {
        "上装": "top",
        "下装": "bottoms",
        "连衣裙": "dress",
        "鞋": "shoes",
        "鞋类": "shoes",
        "配件": "accessory",
    }
    # 大类（cat_l1）最高优先级：配件→accessory、鞋类→shoes。
    # 服装需进一步区分上下装/连衣裙，沿用下方逻辑。
    # 必须早于标题/上下装兜底，否则"测试鞋发售勿拍"等标题里的"鞋"字
    # 会把袜子/帽/包等配件误判为 shoes，"配饰"与"配件"一字之差也会让
    # up_down 兜底 miss。
    if ct == "配件":
        return "accessory"
    if ct == "鞋类":
        return "shoes"
    # 强信号优先于 up_down：连衣裙/连体装是全身款、帽/围巾/手套是配件，
    # 无论源表 up_down 标什么都不应归到 top（修正网球连衣裙被标"上装"等误判）。
    # 帽类用 cat 精确 + 关键词，不能只判 "帽" 子串——会误中"连帽卫衣"。
    # dress 固定只指连衣裙/连体装：不再把 cat「裙装」纳入——「裙装」多为半身裙，
    # 应由 up_down=下装 → bottoms；真连衣裙由 title「连衣裙」命中此强信号保护。
    if "连衣裙" in ttl or "连体" in ttl or cat in ("连衣裙", "连体装"):
        return "dress"
    if (
        cat == "帽类" or "围巾" in cat or "手套" in cat
        or "围巾" in ttl or any(h in ttl for h in ("棒球帽", "鸭舌帽", "遮阳帽", "空顶帽", "渔夫帽", "针织帽", "毛线帽", "网球帽"))
    ):
        return "accessory"
    if ud in ud_map:
        return ud_map[ud]
    if any(x in ct for x in ("鞋", "靴")) or "鞋" in cat or "鞋" in ttl:
        return "shoes"
    if "连衣裙" in cat or "连衣裙" in ttl:
        return "dress"
    if any(x in cat for x in ("帽", "包", "袜")):
        return "accessory"
    if "下装" in ud or "裤" in cat or ("裙" in cat and "连衣" not in cat):
        return "bottoms"
    if "上装" in ud or "外套" in cat or "夹克" in cat or "卫衣" in cat:
        return "top"
    if "上" in ud and "下" not in ud:
        return "top"
    # 小类（cat_l3）兜底：套装/两件套等 up_down=「套装」落不进 ud_map 时，
    # 由小类细分单品补判角色——「梭织上衣」→top、「梭织裤」→bottoms。
    # 仅在强信号均未命中时启用，避免覆盖连衣裙/配件等明确判定；不匹配裸「裙」
    # 以免把「连衣裙/半身裙」错归（连衣裙已由上方强信号判 dress）。
    sc = text_or_empty(sub_cat)
    if sc:
        if "连衣" in sc or "连体" in sc:
            return "dress"
        if any(k in sc for k in (
            "上衣", "上装", "T恤", "POLO", "卫衣", "外套", "夹克", "茄克",
            "衬衫", "衬衣", "毛衣", "背心", "套头", "针织衫", "开衫",
            "棉服", "羽绒", "冲锋衣", "防晒服", "抓绒", "摇粒绒", "马甲",
        )):
            return "top"
        if "裤" in sc or "半裙" in sc or "半身裙" in sc or "裤裙" in sc or "裙裤" in sc:
            return "bottoms"
    return "unknown"


def fila_search_keywords(
    pm: dict[str, str],
    ext: dict[str, str],
    sku_id: str,
) -> str:
    st = text_or_empty(pm.get("search_title"))
    if st:
        return st
    kw = text_or_empty(pm.get("keyword"))
    if kw:
        return kw
    season = first_non_empty(ext.get("season"), ext.get("pro_season"))
    season_normalized = ",".join(normalize_season(season)) if season else ""
    bits = [
        text_or_empty(pm.get("id_alias")),
        text_or_empty(pm.get("pro_title")),
        text_or_empty(pm.get("pro_name")),
        text_or_empty(ext.get("sex")),
        text_or_empty(ext.get("series")),
        season_normalized,
        text_or_empty(ext.get("cat_alias")),
        text_or_empty(ext.get("middle_class")),
        text_or_empty(ext.get("cat_type")),
        text_or_empty(ext.get("up_down")),
        sku_id,
        text_or_empty(ext.get("applicable_scenario")),
        text_or_empty(ext.get("functional_tag")),
    ]
    return ",".join(b for b in bits if b)


def build_search_text(
    *,
    gender: list[str] | str,
    series: str,
    title: str,
    color_name: str,
    role: str,
    spu_id: str,
    sku_id: str,
    extra: str = "",
) -> str:
    if isinstance(gender, list):
        gender_str = " ".join(g for g in gender if g)
    else:
        gender_str = str(gender or "")
    return " ".join(
        x
        for x in [
            BRAND,
            gender_str,
            series,
            title,
            color_name,
            role,
            spu_id,
            sku_id,
            extra,
        ]
        if x
    ).strip()


class EtlLogger:
    """写入 data/logs/etl/{prefix}_{run_id}.jsonl。"""

    def __init__(self, prefix: str, run_id: str | None = None) -> None:
        self.run_id = run_id or datetime.now(timezone.utc).astimezone().strftime(
            "offline_%Y%m%d_%H%M%S",
        )
        etl_dir = logs_dir() / "etl"
        etl_dir.mkdir(parents=True, exist_ok=True)
        self.path = etl_dir / f"{prefix}_{self.run_id}.jsonl"
        self._fh = self.path.open("w", encoding="utf-8")

    def emit(self, event: str, payload: dict[str, Any]) -> None:
        row = {
            "ts": datetime.now(timezone.utc).astimezone().isoformat(),
            "event": event,
            "payload": {**payload, "run_id": self.run_id},
        }
        self._fh.write(json.dumps(row, ensure_ascii=False) + "\n")

    def close(self) -> None:
        self._fh.close()


@dataclass
class ProductTables:
    """FILA 商品表（无 wear_match；搭配来自微导购 + cc_material_product）。"""

    product_dir: Path
    masters: dict[int, dict[str, str]] = field(default_factory=dict)
    exts: dict[int, dict[str, str]] = field(default_factory=dict)
    alias_to_gid: dict[str, int] = field(default_factory=dict)
    attr_by_goods_alias: dict[tuple[int, str], dict[str, str]] = field(
        default_factory=dict,
    )
    sku_id_to_gid: dict[str, int] = field(default_factory=dict)
    color_attrs_by_goods: DefaultDict[int, list[dict[str, str]]] = field(
        default_factory=lambda: defaultdict(list),
    )
    skus_by_goods: DefaultDict[int, list[dict[str, str]]] = field(
        default_factory=lambda: defaultdict(list),
    )
    up_down_by_sku: dict[str, str] = field(default_factory=dict)
    v2_price_by_sku: dict[str, float] = field(default_factory=dict)
    v2_occasion_by_sku: dict[str, list[str]] = field(default_factory=dict)
    v2_style_by_sku: dict[str, list[str]] = field(default_factory=dict)
    v2_cat_l2_by_sku: dict[str, str] = field(default_factory=dict)
    v2_cat_l3_by_sku: dict[str, str] = field(default_factory=dict)
    group_brand_by_sku: dict[str, str] = field(default_factory=dict)
    ext_by_alias: dict[str, dict[str, str]] = field(default_factory=dict)

    @classmethod
    def load(cls, product_dir: Path | None = None) -> ProductTables:
        prod = product_dir or product_dir_path()
        self = cls(product_dir=prod)
        self._load_masters()
        self._load_exts()
        self._load_attrs()
        self._load_product_sku()
        self._load_v2_enrichment()
        return self

    def _load_masters(self) -> None:
        for row in load_csv(self.product_dir / "product_master.csv"):
            gid = norm_id(row.get("id_goods"))
            if gid is None:
                continue
            self.masters[gid] = row
            alias = text_or_empty(row.get("id_alias"))
            if alias and alias not in self.alias_to_gid:
                self.alias_to_gid[alias] = gid

    def _load_exts(self) -> None:
        for row in load_csv(self.product_dir / "product_master_ext.csv"):
            gid = norm_id(row.get("id_goods"))
            if gid is not None:
                self.exts[gid] = row
            alias = text_or_empty(row.get("id_alias"))
            if alias:
                self.ext_by_alias[alias] = row

    def _load_attrs(self) -> None:
        canonical_gids = set(self.alias_to_gid.values())
        for row in load_csv(self.product_dir / "product_attr.csv"):
            gid = norm_id(row.get("id_goods"))
            if gid is None:
                continue
            alias = text_or_empty(row.get("attr_alias"))
            if alias:
                self.attr_by_goods_alias[(gid, alias)] = row
                existing = self.sku_id_to_gid.get(alias)
                if existing is None:
                    self.sku_id_to_gid[alias] = gid
                elif gid in canonical_gids and existing not in canonical_gids:
                    self.sku_id_to_gid[alias] = gid
            if text_or_empty(row.get("id_pac")) != "1":
                continue
            if text_or_empty(row.get("status", "0")) != "0":
                continue
            self.color_attrs_by_goods[gid].append(row)
        for gid in self.color_attrs_by_goods:
            self.color_attrs_by_goods[gid].sort(
                key=lambda x: (
                    norm_id(x.get("order_id")) or 0,
                    text_or_empty(x.get("attr_alias")),
                ),
            )

    def _load_product_sku(self) -> None:
        for row in load_csv(self.product_dir / "product_sku.csv"):
            gid = norm_id(row.get("id_goods"))
            if gid is not None:
                self.skus_by_goods[gid].append(row)

    def _load_v2_enrichment(self) -> None:
        try:
            import product_index_gallery as pig
        except ImportError:
            return
        brief = self.product_dir / "fila_products_brief_prod.xlsx"
        if brief.is_file():
            self.up_down_by_sku.update(pig.load_sku_up_down_map(str(brief)))
        for name in ("fila_products_brief_prod.xlsx",):
            path = self.product_dir / name
            if not path.is_file():
                continue
            rows = pig._load_xlsx_rows(str(path))
            for row in rows:
                sku = pig._row_get(row, "货号", "attrAlias").strip()
                if not sku:
                    continue
                if sku not in self.up_down_by_sku:
                    ud = pig._row_get(row, "上下装", "up_down", "upDown").strip()
                    if ud:
                        self.up_down_by_sku[sku] = ud
                price_raw = pig._row_get(
                    row,
                    "订货会零售价",
                    "price",
                ).strip()
                if price_raw and sku not in self.v2_price_by_sku:
                    try:
                        self.v2_price_by_sku[sku] = float(price_raw)
                    except ValueError:
                        pass
                occ = pig._row_get(row, "场景", "applicable_scenario").strip()
                if occ:
                    self.v2_occasion_by_sku[sku] = split_tags(occ)
                sty = pig._row_get(row, "风格", "style").strip()
                if sty:
                    self.v2_style_by_sku[sku] = split_tags(sty)
                l2 = pig._row_get(row, "中类", "middle_class").strip()
                if l2:
                    self.v2_cat_l2_by_sku[sku] = l2
                l3 = pig._row_get(row, "小类", "cat").strip()
                if l3:
                    self.v2_cat_l3_by_sku[sku] = l3
                gb = pig._row_get(row, "集团品牌", "FASHGRD", "fashgrd").strip()
                if gb:
                    self.group_brand_by_sku[sku] = gb

    def row_score(self, article: str, row: dict[str, str]) -> tuple[int, int, int, str]:
        al = text_or_empty(row.get("attr_alias"))
        g = norm_id(row.get("id_goods")) or 0
        pm = self.masters.get(g, {})
        mid = text_or_empty(pm.get("id_alias"))
        exact = 0 if al == article else 1
        master_ok = 0 if mid == article else 1
        oid = norm_id(row.get("order_id")) or 0
        return (exact, master_ok, oid, al)

    def resolve_article(self, article: str) -> tuple[int | None, str]:
        a = text_or_empty(article)
        if not a:
            return None, ""
        gid = self.alias_to_gid.get(a)
        if gid is not None:
            rows = self.color_attrs_by_goods.get(gid, [])
            pref = [r for r in rows if text_or_empty(r.get("attr_alias")).startswith(a)]
            pool = pref if pref else rows
            if not pool:
                return gid, a
            best = min(pool, key=lambda r: self.row_score(a, r))
            return gid, text_or_empty(best.get("attr_alias")) or a
        cands: list[dict[str, str]] = []
        for rows in self.color_attrs_by_goods.values():
            for r in rows:
                al = text_or_empty(r.get("attr_alias"))
                if al and (al == a or al.startswith(a)):
                    cands.append(r)
        if not cands:
            return None, ""
        best = min(cands, key=lambda r: self.row_score(a, r))
        gid = norm_id(best.get("id_goods"))
        if gid is None:
            return None, ""
        return gid, text_or_empty(best.get("attr_alias")) or a

    def resolve_spu_to_sku(self, spu: str) -> tuple[int | None, str]:
        return self.resolve_article(spu)

    def pick_sku_row(self, gid: int, sku_id: str) -> dict[str, str] | None:
        rows = self.skus_by_goods.get(gid, [])
        if not rows:
            return None
        pa = None
        attr = self.attr_by_goods_alias.get((gid, sku_id))
        if attr:
            pa = text_or_empty(attr.get("id_pa"))
        if pa:
            for row in rows:
                idpas = [
                    p.strip()
                    for p in text_or_empty(row.get("idpas")).split(",")
                    if p.strip()
                ]
                if pa in idpas:
                    return row
        for row in rows:
            if sku_id in text_or_empty(row.get("image")):
                return row
        return rows[0] if rows else None

    def resolve_id_pa(self, gid: int, sku_id: str) -> int | None:
        attr = self.attr_by_goods_alias.get((gid, sku_id))
        if attr:
            pa = norm_id(attr.get("id_pa"))
            if pa is not None:
                return pa
        sku_row = self.pick_sku_row(gid, sku_id)
        if not sku_row:
            return None
        idpas = text_or_empty(sku_row.get("idpas"))
        tokens = [x for x in idpas.split(",") if x.strip().isdigit()]
        pa_ok = {
            str(a.get("id_pa"))
            for a in self.color_attrs_by_goods.get(gid, [])
        }
        for t in tokens:
            if t in pa_ok:
                return int(t)
        if tokens:
            return int(tokens[0])
        return None

    def iter_onsell_goods_ids(self, skip_up_time: bool = False) -> set[int]:
        """product_master 中 onsell∈{1,2} 且 up_time >= 2023-01-01 的 id_goods
        （在售且新上架款；两字段均不能为空）。

        skip_up_time=True 时仅过滤 onsell，不再校验 up_time（用于调试/全量重建）。
        """
        if skip_up_time:
            return self.iter_onsell_only_goods_ids()
        return {
            gid
            for gid, row in self.masters.items()
            if is_onsell(row.get("onsell")) and is_up_time_ok(row.get("up_time"))
        }

    def iter_needed_sku_ids(self, skip_up_time: bool = False) -> Iterator[str]:
        """在售 SKU：product_attr 颜色货号（id_pac=1、status=0）且款在售。

        skip_up_time=True 时不再校验 up_time >= 2023-01-01（仍保留 onsell 与
        老款/电商款过滤），用于调试或全量重建场景。
        跳过老款/电商款 sku_id（数字开头，如 162217109-3），统一为新款字母开头货号。
        """
        onsell = self.iter_onsell_goods_ids(skip_up_time=skip_up_time)
        seen: set[str] = set()
        for gid in sorted(onsell):
            for row in self.color_attrs_by_goods.get(gid, []):
                sku = text_or_empty(row.get("attr_alias"))
                if not sku or sku in seen:
                    continue
                if is_legacy_sku_id(sku):
                    continue
                seen.add(sku)
                yield sku

    def iter_onsell_only_goods_ids(self) -> set[int]:
        """product_master 中 onsell∈{1,2} 的 id_goods（不检查 up_time）。"""
        return {
            gid
            for gid, row in self.masters.items()
            if is_onsell(row.get("onsell"))
        }

    def iter_outfit_article_nos(self) -> set[str]:
        """从 cc_material_product.csv 提取所有搭配引用的 article_no。"""
        cc_path = self.product_dir / "cc_material_product.csv"
        if not cc_path.is_file():
            return set()
        articles: set[str] = set()
        for row in load_csv(cc_path):
            a = text_or_empty(row.get("article_no"))
            if a:
                articles.add(a)
        return articles

    # ── 从 article_no 中提取标准 SPU 的正则 ─────────────────────────
    _RE_ARTICLE_NO = re.compile(r"[A-Z]\d{2}[A-Z]\S+")

    def iter_dphs_outfit_article_nos(self) -> set[str]:
        """从 dphs_outfits.xlsx 的 skus 列提取所有搭配引用的 article_no。"""
        xlsx_path = self.product_dir / "dphs_outfits.xlsx"
        if not xlsx_path.is_file():
            return set()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            sku_col = headers.index("skus") if "skus" in headers else -1
            if sku_col < 0:
                wb.close()
                return set()
            articles: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw = str(row[sku_col] or "").strip()
                if not raw:
                    continue
                for part in raw.split(","):
                    a = part.strip()
                    if a:
                        articles.add(a)
            wb.close()
            return articles
        except Exception:
            return set()

    def iter_outfits_unique_article_nos(self) -> set[str]:
        """从 outfits_unique.txt 提取所有搭配引用的 article_no。

        每行格式: 类别SKU-类别SKU-...，如 上衣T11U613701FDP-内搭T11W611103FBK
        """
        txt_path = self.product_dir / "outfits_unique.txt"
        if not txt_path.is_file():
            return set()
        articles: set[str] = set()
        with txt_path.open(encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                for segment in line.split("-"):
                    m = self._RE_ARTICLE_NO.search(segment)
                    if m:
                        articles.add(m.group(0))
        return articles

    def _ensure_alias_prefix_index(self) -> None:
        """Lazily build an alias prefix index for fast article_no lookups."""
        if hasattr(self, "_alias_prefix_index"):
            return
        idx: dict[str, int] = {}
        for gid, rows in self.color_attrs_by_goods.items():
            for r in rows:
                alias = text_or_empty(r.get("attr_alias"))
                if alias and alias not in idx:
                    idx[alias] = gid
        self._alias_prefix_index: dict[str, int] = idx

    def resolve_article_to_sku_ids(self, article_no: str) -> list[str]:
        """article_no → sku_id 列表，复用 alias_to_gid + color_attrs_by_goods。"""
        a = (article_no or "").strip()
        if not a:
            return []
        gid = self.alias_to_gid.get(a) or self.sku_id_to_gid.get(a)
        if gid is not None:
            rows = self.color_attrs_by_goods.get(gid, [])
            pref = [
                text_or_empty(r.get("attr_alias"))
                for r in rows
                if text_or_empty(r.get("attr_alias")).startswith(a)
            ]
            if pref:
                return [s for s in pref if s]
            all_aliases = [text_or_empty(r.get("attr_alias")) for r in rows]
            return [s for s in all_aliases if s] or [a]
        # Fast path: use prefix index instead of O(N) full scan
        self._ensure_alias_prefix_index()
        for alias, g in self._alias_prefix_index.items():
            if alias.startswith(a):
                return [alias]
        return []

    # 标准 SPU 模式: 品牌(1)+季(2)+性别(1)+品类(5-6)+系列标识(1), 如 F11M348208F
    _RE_STD_SPU = re.compile(r"[A-Z]\d{2}[A-Z]\d{5,6}[A-Z]")

    def _resolve_ext_by_alias(self, spu: str) -> dict[str, str]:
        """通过 SPU alias 查找 ext，支持规范化回退。

        product_master 中 id_alias 可能带后缀数字(F11M348208F1)或
        前缀年份(2026F13M618311F)，而 product_master_ext 用原始 alias。
        先收集所有候选 alias，返回信息最丰富的（优先有 season/sex 的记录）。
        """
        candidates: list[str] = [spu]
        # 去除末尾数字后缀 (F11M348208F1 -> F11M348208F)
        stripped = re.sub(r"\d+$", "", spu)
        if stripped and stripped != spu:
            candidates.append(stripped)
        # 去除开头4位年份前缀 (2026F13M618311F -> F13M618311F)
        if len(spu) > 4 and spu[:4].isdigit():
            candidates.append(spu[4:])
        # 提取标准 SPU 模式
        m = self._RE_STD_SPU.search(spu)
        if m and m.group(0) != spu:
            candidates.append(m.group(0))
        best: dict[str, str] = {}
        best_score = -1
        for c in candidates:
            hit = self.ext_by_alias.get(c)
            if not hit:
                continue
            score = bool(hit.get("sex")) + bool(hit.get("season") or hit.get("pro_season"))
            if score > best_score:
                best = hit
                best_score = score
                if score >= 2:
                    break
        return best

    def compute_max_update_times(self) -> dict[int, datetime]:
        """Compute the latest update timestamp per goods_id from source tables.

        Scans product_master.updated_at and product_master_ext.update_time.
        Returns {goods_id: max_datetime}; goods with no timestamp are omitted.
        """
        result: dict[int, datetime] = {}
        for gid, row in self.masters.items():
            dt = _parse_iso_datetime(row.get('updated_at'))
            if dt is not None:
                result[gid] = dt
        for gid, row in self.exts.items():
            dt = _parse_iso_datetime(row.get('update_time'))
            if dt is not None:
                if gid not in result or dt > result[gid]:
                    result[gid] = dt
        return result

    def build_sku_record(self, sku_id: str) -> dict[str, Any] | None:
        gid = self.sku_id_to_gid.get(sku_id)
        if gid is None:
            for g, rows in self.color_attrs_by_goods.items():
                for r in rows:
                    if text_or_empty(r.get("attr_alias")) == sku_id:
                        gid = g
                        break
                if gid is not None:
                    break
        if gid is None or gid not in self.masters:
            return None
        master = self.masters[gid]
        ext = self.exts.get(gid, {})
        spu = normalize_spu_id(text_or_empty(master.get("id_alias")))
        if not ext and spu:
            ext = self._resolve_ext_by_alias(spu)
        sku_row = self.pick_sku_row(gid, sku_id) or {}
        attr = self.attr_by_goods_alias.get((gid, sku_id), {})
        title = first_non_empty(
            master.get("pro_title"),
            master.get("pro_name"),
        ) or ""
        up_down = self.up_down_by_sku.get(
            sku_id,
            text_or_empty(ext.get("up_down")),
        )
        up_down = _SX_UP_DOWN_MAP.get(up_down, up_down)
        # 优先 cat_alias(ext) 再 xlsx 中类：xlsx "中类"列口径粗（短T类/包类/下装类…），
        # 优先级过高会盖掉 cat_alias 的细值，故 ext 优先、xlsx 仅作回退。
        cat_l2 = fila_category_l2(ext) or normalize_category_l2(self.v2_cat_l2_by_sku.get(sku_id) or "")
        # cat_l3 提前到 role 推断之前：套装/两件套需由小类细分单品补判角色。
        cat_l3 = self.v2_cat_l3_by_sku.get(sku_id) or text_or_empty(
            ext.get("short_category"),
        )
        if not up_down and cat_l2:
            up_down = CATEGORY_L2_UP_DOWN.get(cat_l2, "")
        if not up_down:
            up_down = infer_up_down_from_title(title)
        role = infer_role(ext, title, up_down, cat_l3)
        id_pa = self.resolve_id_pa(gid, sku_id)
        price = float(
            sku_row.get("shop_price")
            or master.get("price")
            or self.v2_price_by_sku.get(sku_id)
            or 0,
        )
        gender = normalize_gender_to_list(
            text_or_empty(ext.get("sex")),
            title=title,
            sku_id=sku_id,
        )
        age = normalize_age(ext.get("age"))
        if not age:
            # 源 age 列缺失时，从标题兜底推断（只匹配童装年龄段关键词，成人款返回 ""）
            age = infer_age_from_title(title)
        up_time = normalize_up_time(master.get("up_time"))
        if "童" in title and gender:
            upgraded: list[str] = []
            for g in gender:
                if g == "男":
                    upgraded.append("男童")
                elif g == "女":
                    upgraded.append("女童")
                else:
                    upgraded.append(g)
            seen: set[str] = set()
            gender = [x for x in upgraded if not (x in seen or seen.add(x))]
        season_list = normalize_season([
            text_or_empty(ext.get("season")),
            text_or_empty(ext.get("pro_season")),
        ])
        if not season_list and spu:
            ext_alias = self._resolve_ext_by_alias(spu)
            if ext_alias and ext_alias is not ext:
                season_list = normalize_season([
                    text_or_empty(ext_alias.get("season")),
                    text_or_empty(ext_alias.get("pro_season")),
                ])
        color_name = extract_color_name(
            text_or_empty(sku_row.get("attr_name"))
            or text_or_empty(attr.get("attr_name")),
        )
        if not color_name:
            color_name = text_or_empty(attr.get("attr_name"))
        occ = self.v2_occasion_by_sku.get(sku_id) or split_tags(
            text_or_empty(ext.get("applicable_scenario")),
        )
        cat_l1 = text_or_empty(ext.get("cat_type"))
        sty = self.v2_style_by_sku.get(sku_id) or split_tags(
            text_or_empty(ext.get("functional_tag")),
        )
        search_text = build_search_text(
            gender=gender,
            series=text_or_empty(ext.get("series")),
            title=title,
            color_name=color_name,
            role=role,
            spu_id=spu,
            sku_id=sku_id,
        )
        search_keywords = fila_search_keywords(master, ext, sku_id)
        descent_extra = build_descent_extra_fields(
            master,
            ext,
            self.color_attrs_by_goods.get(gid, []),
            len(self.skus_by_goods.get(gid, [])),
        )
        return {
            "sku_id": sku_id,
            "spu_id": spu,
            "id_goods": gid,
            "id_pa": id_pa,
            "title": title,
            "brand": BRAND,
            "group_brand": self.group_brand_by_sku.get(sku_id, ""),
            "gender": gender,
            "age": age,
            "up_time": up_time,
            "category_l1": cat_l1,
            "category_l2": cat_l2,
            "category_l3": cat_l3,
            "role": role,
            "up_down_raw": up_down,
            "series": text_or_empty(ext.get("series")),
            "sub_series": text_or_empty(ext.get("cat")),
            "season": season_list,
            "occasion_tags": occ,
            "style_tags": sty,
            "color_name": color_name,
            "color_family": color_name,
            "color_series": map_color_to_series_list(color_name),
            "price": price,
            "fabric_function": split_tags(
                text_or_empty(ext.get("fabric_function")),
            ),
            "material": first_non_empty(ext.get("material"), ext.get("fabric"))
            or "",
            "display_image": "",
            "index_images": [],
            "tryon_image": "",
            "all_images": [],
            "image_quality": {
                "display_score": 0.0,
                "index_score": 0.0,
                "tryon_score": 0.0,
                "is_tryon_ready": False,
            },
            "search_text": search_text,
            "search_keywords": search_keywords,
            "attr_name": color_name,
            "excluded_from_pairing": is_excluded_from_pairing(cat_l2),
            "layer": extract_layer(cat_l2, title),
            "coverage": extract_coverage(role, cat_l2, title),
            "length_class": resolve_length_class(role, cat_l2, title, sku_id, search_keywords, cat_l3),
            "is_intimate": extract_is_intimate(cat_l2, title),
            "modeling": normalize_modeling(ext.get("modeling")),
            "scene_domain": extract_scene_domain(
                cat_l1, cat_l2, role, occ, title, search_keywords,
                text_or_empty(ext.get("series")),
                text_or_empty(ext.get("cat")),
            ),
            **descent_extra,
        }


def product_dir_path() -> Path:
    return product_dir()


def load_vlm_index(product_dir: Path) -> dict[str, dict]:
    """sku_id -> {tryon_url, index_images} from fila_sku_selected_images.csv。

    返回结构::

        {
            "SKU001": {
                "tryon_url": "https://...",
                "index_images": ["https://...", "https://..."],
            },
            ...
        }
    """
    out: dict[str, dict] = {}
    import json as _json
    for name in ("fila_sku_selected_images.csv",):
        path = product_dir / name
        if not path.is_file():
            continue
        for row in load_csv(path):
            sku = text_or_empty(
                row.get("attr_alias")
                or row.get("sku_id")
                or row.get("货号"),
            )
            if not sku:
                continue
            # tryon_image: 取 white_front_url 兜底
            tryon_url = first_non_empty(
                row.get("white_front_url"),
                row.get("tryon_image"),
                row.get("selected_url"),
                row.get("image_url"),
                row.get("path"),
            ) or ""
            # index_images: JSON 数组列
            index_images: list[str] = []
            raw_idx = (row.get("index_images") or "").strip()
            if raw_idx and raw_idx != "[]":
                try:
                    parsed = _json.loads(raw_idx)
                    if isinstance(parsed, list):
                        index_images = [
                            str(u).strip() for u in parsed if str(u).strip()
                        ]
                except (ValueError, TypeError):
                    pass
            if sku and (tryon_url or index_images):
                out[sku] = {
                    "tryon_url": tryon_url,
                    "index_images": index_images,
                }
    return out


def load_vlm_excluded_skus(product_dir: Path) -> set[str]:
    """VLM 跑过且有候选图、但全部被判为非商品图（吊牌/水洗标/尺码表等）的货号集合。

    判据：``candidate_count>0`` 且 ``white_front_url`` 为空 且 ``index_images``
    为空。这些货号本就没有可用主图，``select_images`` 不得用 fallback 把吊牌图
    当 master 选回 index_images / tryon_image（否则会被占位图过滤放行入库）。

    fila_sku_selected_images.csv 不存在时返回空集合（未跑 VLM 时原逻辑不变）。
    """
    out: set[str] = set()
    import json as _json
    path = product_dir / "fila_sku_selected_images.csv"
    if not path.is_file():
        return out
    for row in load_csv(path):
        sku = text_or_empty(
            row.get("attr_alias")
            or row.get("sku_id")
            or row.get("货号"),
        )
        if not sku:
            continue
        try:
            cand = int(float(row.get("candidate_count") or 0))
        except (TypeError, ValueError):
            cand = 0
        if cand <= 0:
            continue
        white = first_non_empty(
            row.get("white_front_url"),
            row.get("selected_url"),
        )
        raw_idx = (row.get("index_images") or "").strip()
        has_idx = bool(raw_idx) and raw_idx != "[]"
        if has_idx:
            try:
                parsed = _json.loads(raw_idx)
                has_idx = isinstance(parsed, list) and any(
                    str(u).strip() for u in parsed
                )
            except (ValueError, TypeError):
                has_idx = False
        if not white and not has_idx:
            out.add(sku)
    return out


@lru_cache(maxsize=1)
def load_length_class_vlm_index() -> dict[str, str]:
    """sku_id -> length_class，来自 VLM 离线判定 CSV（data/processed/sku_length_vlm.csv）。

    仅收录**成功且非 n/a** 的记录（length_class ∈ {long, short}，error 为空）。
    用于在 ETL 构建时回补 ``extract_length_class`` 落 n/a 的上装/下装：规则推导为
    n/a 时回退查本表，命中则用 VLM 值覆盖。n/a 的 VLM 记录不收录（无值可补）。
    CSV 不存在时返回空 dict（未跑 VLM 时不影响原逻辑）。
    """
    path = processed_dir() / "sku_length_vlm.csv"
    if not path.is_file():
        return {}
    out: dict[str, str] = {}
    for row in load_csv(path):
        if row.get("error"):
            continue
        lc = text_or_empty(row.get("length_class")).strip().lower()
        if lc not in ("long", "short"):
            continue
        sku = text_or_empty(row.get("sku_id"))
        if sku:
            out[sku] = lc
    return out


def resolve_length_class(
    role: str, cat_l2: str, title: str, sku_id: str,
    search_keywords: str = "", cat_l3: str = "",
) -> str:
    """规则推导 length_class；为 n/a 时回退查 VLM CSV（仅补 n/a，不覆盖已有 short/long）。

    search_keywords 作为 title 之外的兜底文本喂给 ``extract_length_class``：当
    cat_l2 被中类合并抹掉长短信息（如「梭织五分裤」→「梭织裤」）而 title 又无
    长短关键词时，search_keywords 里的「五分裤/短裤/长裤」可补判 short/long。
    cat_l3（小类）参与下装 rule1/rule2 的「title/中类/小类」扫描。

    VLM 回退仅对 role∈{top,bottoms} 生效——连衣裙/连体装等全身款本就 n/a，
    不应套用当初误判为 top 时跑出的 VLM 长短款。

    泳装类同样跳过 VLM 回退：length 在沙滩域不作 season 代理，长款防晒开衫×
    短款泳裤是正确搭配，VLM 会按视觉长短判 long/short 反而重新触发季节冲突。
    """
    lc = extract_length_class(role, cat_l2, title, search_keywords, cat_l3)
    if (
        lc == "n/a"
        and (role or "").strip().lower() in ("top", "bottoms")
        and not is_swimwear(cat_l2, title)
    ):
        vlm = load_length_class_vlm_index().get(sku_id)
        if vlm in ("long", "short"):
            return vlm
    return lc


def load_ai_select_index(
    product_dir: Path,
) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    """货号 / 款号 -> VLM 白底图选型元数据（供详情页 ai_select 展示）。"""
    by_sku: dict[str, dict[str, str]] = {}
    by_spu: dict[str, dict[str, str]] = {}
    path = product_dir / "fila_sku_selected_images.csv"
    if not path.is_file():
        return by_sku, by_spu
    for row in load_csv(path):
        url = first_non_empty(
            row.get("white_front_url"),
            row.get("selected_url"),
        )
        if not url:
            continue
        sku = text_or_empty(
            row.get("货号") or row.get("attr_alias") or row.get("sku_id"),
        )
        spu = text_or_empty(
            row.get("款号") or row.get("id_alias") or row.get("spu_id"),
        )
        note = first_non_empty(row.get("tryon_reason"), row.get("model_reason"), row.get("note"))
        payload = {
            "path": url,
            "note": note,
            "candidate_count": text_or_empty(row.get("candidate_count")),
            "chosen_id_pa": text_or_empty(row.get("chosen_id_pa")),
            "chosen_order_id": text_or_empty(row.get("chosen_order_id")),
            "chosen_image_type": (
                text_or_empty(row.get("chosen_image_type")) or "ai_select"
            ),
        }
        if sku:
            by_sku[sku] = payload
        if spu:
            by_spu[spu] = payload
    return by_sku, by_spu


def outfit_completeness(roles: set[str]) -> float:
    core = {"top", "bottoms", "shoes", "dress"}
    hit = len(roles & core)
    if "dress" in roles and "shoes" in roles:
        return 1.0
    return round(min(1.0, hit / 3.0), 3)


def outfit_quality_score(
    completeness: float,
    tryon_coverage: float,
    n_items: int,
) -> float:
    base = 0.25 + 0.35 * completeness + 0.25 * tryon_coverage
    if n_items >= 3:
        base += 0.05
    return round(min(1.0, base), 3)
