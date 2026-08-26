"""搭配话术库：从 dphs_outfits.xlsx 加载 tag→reason 映射，支持标签匹配召回话术。"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from backend.config import get_root, load_config

logger = logging.getLogger(__name__)

# 单例缓存
_STORE: DphsReasonStore | None = None


class DphsReasonStore:
    """话术库：tag→reason 倒排索引 + 子串匹配。"""

    def __init__(self) -> None:
        # 每条记录：{"tags": ["标签A", "标签B"], "reason": "话术原文"}
        self.entries: list[dict[str, Any]] = []
        # tag→[entry index] 倒排索引
        self._tag_index: dict[str, list[int]] = {}
        self._loaded = False

    def load(self, xlsx_path: Path | None = None) -> None:
        if self._loaded:
            return
        if xlsx_path is None:
            root = get_root()
            cfg = load_config()
            table_dir = (cfg.get("paths") or {}).get("product_dir", "data/tables")
            xlsx_path = root / table_dir / "dphs_outfits.xlsx"
        if not xlsx_path.is_file():
            logger.warning("dphs_outfits.xlsx not found: %s", xlsx_path)
            self._loaded = True
            return
        self._parse_xlsx(xlsx_path)
        self._loaded = True
        logger.info(
            "DphsReasonStore loaded: %d entries, %d unique tags",
            len(self.entries),
            len(self._tag_index),
        )

    def _parse_xlsx(self, path: Path) -> None:
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed, cannot load dphs_outfits.xlsx")
            return
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            tags_idx = headers.index("tags")
            reason_idx = headers.index("reason")
        except ValueError:
            logger.error("dphs_outfits.xlsx missing 'tags' or 'reason' column, headers=%s", headers)
            wb.close()
            return
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_tags = str(row[tags_idx] or "").strip()
            raw_reason = str(row[reason_idx] or "").strip()
            if not raw_tags or not raw_reason:
                continue
            tags = [t.strip() for t in raw_tags.split("#") if t.strip()]
            entry_idx = len(self.entries)
            self.entries.append({"tags": tags, "reason": raw_reason})
            for tag in tags:
                self._tag_index.setdefault(tag, []).append(entry_idx)
        wb.close()

    def match_reasons(self, outfit_text: str, max_results: int = 3) -> list[dict[str, Any]]:
        """根据搭配描述文本，子串匹配命中的话术。

        Args:
            outfit_text: 搭配单品的标题、品类等拼接文本。
            max_results: 最多返回几条话术。

        Returns:
            [{"matched_tags": ["标签A"], "reason": "话术原文"}, ...]
        """
        self.load()
        if not outfit_text or not self._tag_index:
            return []
        text_lower = outfit_text.lower()
        # 记录每条 entry 被命中的标签数和命中标签列表
        hit_map: dict[int, list[str]] = {}
        for tag, entry_indices in self._tag_index.items():
            tag_lower = tag.lower()
            if tag_lower in text_lower:
                for idx in entry_indices:
                    hit_map.setdefault(idx, []).append(tag)
        if not hit_map:
            return []
        # 按命中标签数降序排列
        scored = sorted(hit_map.items(), key=lambda x: -len(x[1]))
        results: list[dict[str, Any]] = []
        for entry_idx, matched_tags in scored[:max_results]:
            entry = self.entries[entry_idx]
            results.append({
                "matched_tags": matched_tags,
                "reason": entry["reason"],
            })
        return results


def get_dphs_reason_store() -> DphsReasonStore:
    """获取话术库单例。"""
    global _STORE
    if _STORE is None:
        _STORE = DphsReasonStore()
        _STORE.load()
    return _STORE


def extract_outfit_text(card: dict[str, Any]) -> str:
    """从搭配卡片中提取用于标签匹配的文本。"""
    parts: list[str] = []
    name = str(card.get("name") or "")
    if name:
        parts.append(name)
    for item in card.get("items") or []:
        title = str(item.get("title") or "")
        if title:
            parts.append(title)
        role = str(item.get("role") or "")
        if role:
            parts.append(role)
    return " ".join(parts)


def match_outfit_reasons(card: dict[str, Any], max_results: int = 3) -> list[dict[str, Any]]:
    """从搭配卡片匹配话术。"""
    text = extract_outfit_text(card)
    return get_dphs_reason_store().match_reasons(text, max_results=max_results)


def format_reasons_as_fewshot(matches: list[dict[str, Any]]) -> str:
    """将匹配到的话术格式化为 few-shot 参考文本。"""
    if not matches:
        return ""
    lines = ["## 参考话术风格（请模仿以下风格和口吻撰写推荐理由，突出卖点和场景感）"]
    for i, m in enumerate(matches, 1):
        tags_str = " ".join(f"#{t}" for t in m["matched_tags"])
        lines.append(f"\n示例{i}：")
        lines.append(f"标签：{tags_str}")
        lines.append(f"话术：{m['reason']}")
    return "\n".join(lines)


def build_template_reason(card: dict[str, Any]) -> str:
    """模板直出模式：用命中话术拼接推荐理由，无命中时返回兜底文案。"""
    matches = match_outfit_reasons(card)
    if not matches:
        return "这套搭配风格统一，单品之间色彩和品类搭配协调，适合日常穿着。"
    # 取命中最多标签的那条话术
    return matches[0]["reason"]
