"""Attach allProductImages (product_image rows) to product_index items."""

from __future__ import annotations

import csv
import os
import shutil
import sys
import tempfile
from typing import Any, Dict, List, Optional, Tuple


def _norm_int(raw: Any) -> Optional[int]:
    if raw is None:
        return None
    s = str(raw).strip().strip('"').lstrip("'").strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _load_csv_rows(path: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with open(path, 'r', encoding='utf-8-sig', newline='') as handle:
        for row in csv.DictReader(handle):
            rows.append({k: (v or '') for k, v in row.items()})
    return rows


def _load_xlsx_rows(path: str) -> List[Dict[str, str]]:
    try:
        import openpyxl
    except ImportError:
        print(
            'product_index_gallery: openpyxl not installed, skip xlsx:',
            path,
            file=sys.stderr,
        )
        return []
    tmp_copy = None
    wb = None
    try:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            low = str(exc).lower()
            if 'xls' in low and 'support' in low:
                fd, tmp_copy = tempfile.mkstemp(suffix='.xlsx')
                os.close(fd)
                shutil.copyfile(path, tmp_copy)
                wb = openpyxl.load_workbook(
                    tmp_copy,
                    read_only=True,
                    data_only=True,
                )
            else:
                print(
                    f'product_index_gallery: skip spreadsheet {path!r}: {exc}',
                    file=sys.stderr,
                )
                return []
        ws = wb.active
        header_row = next(
            ws.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if not header_row:
            return []
        headers = [str(c or '').strip() for c in header_row]
        out: List[Dict[str, str]] = []
        for tup in ws.iter_rows(min_row=2, values_only=True):
            if not tup:
                continue
            row: Dict[str, str] = {}
            for idx, key in enumerate(headers):
                if not key:
                    continue
                val = tup[idx] if idx < len(tup) else None
                if val is None:
                    row[key] = ''
                else:
                    row[key] = str(val).strip()
            if any(v for v in row.values()):
                out.append(row)
        return out
    except Exception as exc:
        print(
            f'product_index_gallery: skip spreadsheet {path!r}: {exc}',
            file=sys.stderr,
        )
        return []
    finally:
        if wb is not None:
            wb.close()
        if tmp_copy and os.path.isfile(tmp_copy):
            try:
                os.unlink(tmp_copy)
            except OSError:
                pass


def _load_xls_rows(path: str) -> List[Dict[str, str]]:
    try:
        import xlrd
    except ImportError:
        print(
            'product_index_gallery: install xlrd to read .xls product_image:',
            path,
            file=sys.stderr,
        )
        return []
    try:
        book = xlrd.open_workbook(path)
    except Exception as exc:
        low = str(exc).lower()
        if 'xlsx' in low or 'xlsb' in low:
            return []
        print(
            f'product_index_gallery: xlrd failed for {path!r}: {exc}',
            file=sys.stderr,
        )
        return []
    sheet = book.sheet_by_index(0)
    if sheet.nrows < 1:
        return []
    headers = [
        str(sheet.cell_value(0, col)).strip()
        for col in range(sheet.ncols)
    ]
    out: List[Dict[str, str]] = []
    for row_idx in range(1, sheet.nrows):
        row: Dict[str, str] = {}
        for col_idx, key in enumerate(headers):
            if not key:
                continue
            val = sheet.cell_value(row_idx, col_idx)
            if val is None or val == '':
                row[key] = ''
            elif isinstance(val, float) and val == int(val):
                row[key] = str(int(val))
            else:
                row[key] = str(val).strip()
        if any(v for v in row.values()):
            out.append(row)
    return out


def _image_rows_from_file(path: str) -> List[Dict[str, str]]:
    if not path or not os.path.isfile(path):
        return []
    lower = path.lower()
    if lower.endswith('.csv'):
        return _load_csv_rows(path)
    if lower.endswith('.xlsx'):
        return _load_xlsx_rows(path)
    if lower.endswith('.xls'):
        xls_rows = _load_xls_rows(path)
        if xls_rows:
            return xls_rows
        return _load_xlsx_rows(path)
    return []


def _row_get(row: Dict[str, str], *keys: str) -> str:
    for key in keys:
        if key in row and row[key]:
            return row[key]
    lower_map = {k.lower(): v for k, v in row.items()}
    for key in keys:
        v = lower_map.get(key.lower(), '')
        if v:
            return v
    return ''


def load_images_by_goods(path: Optional[str]) -> Dict[int, List[Dict[str, Any]]]:
    """id_goods -> rows with status==1, non-empty path."""
    rows = _image_rows_from_file(path) if path else []
    by_gid: Dict[int, List[Dict[str, Any]]] = {}
    for row in rows:
        if str(_row_get(row, 'status')).strip() != '1':
            continue
        url = _row_get(row, 'path', 'Path', 'url').strip()
        if not url:
            continue
        gid = _norm_int(_row_get(row, 'id_goods', 'idGoods'))
        if gid is None:
            continue
        id_pa = _norm_int(_row_get(row, 'id_pa', 'idPa'))
        if id_pa is None:
            id_pa = 0
        order_id = _norm_int(_row_get(row, 'order_id', 'orderId'))
        if order_id is None:
            order_id = 0
        image_type = _row_get(row, 'image_type', 'imageType').strip()
        by_gid.setdefault(gid, []).append({
            'path': url,
            'idPa': id_pa,
            'orderId': order_id,
            'imageType': image_type,
        })
    return by_gid


def _sort_key(entry: Dict[str, Any]) -> Tuple[str, int, int, str]:
    return (
        str(entry.get('imageType') or ''),
        int(entry.get('orderId') or 0),
        int(entry.get('idPa') or 0),
        str(entry.get('path') or ''),
    )


def _unique_index_items(
    by_alias: Dict[str, Any],
    by_id_goods: Dict[str, Any],
) -> List[Any]:
    seen = set()
    out: List[Any] = []
    for bucket in (by_alias, by_id_goods):
        for item in bucket.values():
            iid = id(item)
            if iid in seen:
                continue
            seen.add(iid)
            out.append(item)
    return out


def load_sku_up_down_map(path: Optional[str]) -> Dict[str, str]:
    """货号 -> 上下装（来自 data/tables/商品_斐乐v2.xlsx 等表）。"""
    rows = _load_xlsx_rows(path) if path and os.path.isfile(path) else []
    out: Dict[str, str] = {}
    for row in rows:
        sku = _row_get(row, '货号', 'attrAlias').strip()
        ud = _row_get(row, '上下装', 'up_down', 'upDown').strip()
        if sku and ud:
            out[sku] = ud
    return out


def _apply_up_down_to_item(
    item: Dict[str, Any],
    by_sku: Dict[str, str],
) -> bool:
    """当 attributes.upDown 为空时，用货号字典补全。有值则不改。"""
    attrs = item.setdefault('attributes', {})
    cur = attrs.get('upDown')
    if cur is not None and str(cur).strip():
        return False
    alias = str(item.get('attrAlias') or '').strip()
    val = by_sku.get(alias)
    if not val:
        return False
    attrs['upDown'] = val
    return True


def spread_up_down_to_outfits(
    outfits: List[Any],
    xlsx_path: Optional[str],
) -> int:
    """为每个搭配条目写入 Excel 中的上下装（仅填补空值）。"""
    by_sku = load_sku_up_down_map(xlsx_path)
    if not by_sku:
        return 0
    n = 0
    for outfit in outfits:
        for item in outfit.get('items') or []:
            if _apply_up_down_to_item(item, by_sku):
                n += 1
    return n


def enrich_up_down(
    by_alias: Dict[str, Any],
    by_id_goods: Dict[str, Any],
    xlsx_path: Optional[str],
) -> int:
    """Mutate items' attributes.upDown from 商品_斐乐v2.xlsx（货号 join）。"""
    by_sku = load_sku_up_down_map(xlsx_path)
    if not by_sku:
        return 0
    n = 0
    for item in _unique_index_items(by_alias, by_id_goods):
        if _apply_up_down_to_item(item, by_sku):
            n += 1
    return n


def load_ai_select_rows(
    path: Optional[str],
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    """Return (by_alias, by_style) for ai_select csv rows.

    Compatible with both old `fila_white_front_by_style.csv`
    and new `fila_sku_selected_images.csv`.
    """
    if not path or not os.path.isfile(path):
        return {}, {}
    rows = _load_csv_rows(path)
    by_alias: Dict[str, Dict[str, Any]] = {}
    by_style: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        url = _row_get(row, 'white_front_url', 'whiteFrontUrl').strip()
        if not url:
            continue
        style = _row_get(row, '款号', 'idAlias').strip()
        alias = _row_get(row, '货号', 'attrAlias').strip()
        note = _row_get(row, 'note').strip()
        cand = _row_get(row, 'candidate_count', 'candidateCount').strip()
        chosen_id_pa = _row_get(row, 'chosen_id_pa', 'chosenIdPa').strip()
        chosen_order_id = _row_get(row, 'chosen_order_id', 'chosenOrderId').strip()
        chosen_image_type = _row_get(row, 'chosen_image_type', 'chosenImageType').strip()
        payload = {
            'path': url,
            'imageType': 'ai_select',
            'note': note,
            'candidateCount': cand,
            'chosenIdPa': chosen_id_pa,
            'chosenOrderId': chosen_order_id,
            'chosenImageType': chosen_image_type,
        }
        if alias:
            by_alias[alias] = dict(payload)
        if style:
            by_style[style] = dict(payload)
    return by_alias, by_style


def _apply_ai_select_to_item(
    item: Dict[str, Any],
    ai_by_alias: Dict[str, Dict[str, Any]],
    ai_by_style: Dict[str, Dict[str, Any]],
) -> bool:
    """Set or clear item['aiSelect']. Returns True if a payload was attached."""
    alias = str(item.get('attrAlias') or '').strip()
    style = str(item.get('idAlias') or '').strip()
    row = ai_by_alias.get(alias) or ai_by_style.get(style)
    if row:
        item['aiSelect'] = dict(row)
        return True
    item.pop('aiSelect', None)
    return False


def spread_ai_select_to_outfits(
    outfits: List[Any],
    csv_path: Optional[str],
) -> int:
    """Attach aiSelect to every item in every outfit (chunk JSON thumbnails)."""
    ai_by_alias, ai_by_style = load_ai_select_rows(csv_path)
    if not ai_by_alias and not ai_by_style:
        for outfit in outfits:
            for item in outfit.get('items') or []:
                item.pop('aiSelect', None)
        return 0
    n = 0
    for outfit in outfits:
        for item in outfit.get('items') or []:
            if _apply_ai_select_to_item(item, ai_by_alias, ai_by_style):
                n += 1
    return n


def enrich_ai_select(
    by_alias: Dict[str, Any],
    by_id_goods: Dict[str, Any],
    csv_path: Optional[str],
) -> int:
    """Mutate items with aiSelect (货号优先、款号兜底)."""
    ai_by_alias, ai_by_style = load_ai_select_rows(csv_path)
    if not ai_by_alias and not ai_by_style:
        for item in _unique_index_items(by_alias, by_id_goods):
            item.pop('aiSelect', None)
        return 0

    n = 0
    for item in _unique_index_items(by_alias, by_id_goods):
        if _apply_ai_select_to_item(item, ai_by_alias, ai_by_style):
            n += 1
    return n


def enrich_product_index(
    by_alias: Dict[str, Any],
    by_id_goods: Dict[str, Any],
    img_path: Optional[str],
) -> int:
    """Mutate items with allProductImages sorted by imageType, orderId, idPa.

    Returns count of items that received a non-empty gallery.
    """
    by_gid = load_images_by_goods(img_path)
    if not by_gid:
        for item in _unique_index_items(by_alias, by_id_goods):
            item['allProductImages'] = []
        return 0

    n_non_empty = 0
    for item in _unique_index_items(by_alias, by_id_goods):
        gid = _norm_int(item.get('idGoods'))
        raw = list(by_gid.get(gid, [])) if gid is not None else []
        raw.sort(key=_sort_key)
        item['allProductImages'] = raw
        if raw:
            n_non_empty += 1
    return n_non_empty
